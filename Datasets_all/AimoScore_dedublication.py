import pickle
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

INPUT_FILE = "AimoScore_WeakLink_big_scores.xls"  # or .xls if using xlrd engine
SCORE_PICKLE = "AimoScore_WeakLink_big_scores.pkl"
OUTPUT_FILE = "AimoScore_deduped.xlsx"
REMOVED_REPORT = "AimoScore_removed_report.csv"

# old implemenation to load from original excel
#sheets = pd.read_excel(INPUT_FILE, sheet_name=None)
#df = next(iter(sheets.values()))
with open(SCORE_PICKLE, "rb") as f:
    sheets = pickle.load(f)
    df = sheets["Sheet1"]
    #print(df)

required_cols = {"AimoScore", "EstimatedScore"}
if not required_cols.issubset(df.columns):
    raise ValueError(f"Input file must contain columns: {required_cols}. Found: {df.columns.tolist()}")

counts = df["AimoScore"].value_counts()
duplicated_values = set(counts[counts > 1].index)
mask_removed = df["AimoScore"].isin(duplicated_values)
removed_rows = df[mask_removed].copy().reset_index(drop=True)
kept_rows = df[~mask_removed].copy()

# Add comparison columns for removed rows
removed_rows["Diff"] = removed_rows["EstimatedScore"] - removed_rows["AimoScore"]


# Convert RelDiff into boolean Quality estimator: True if absolute rel diff > 0.1
removed_rows["Threshold"] = removed_rows["Diff"].apply(
    lambda x: bool(abs(x) > 0.1) if pd.notna(x) else False
)

# Only keep the requested columns for the removed report
removed_report_df = removed_rows[["EstimatedScore", "Diff", "Threshold"]]

# Save removed report (overwrite previous save)
removed_report_df.to_csv(REMOVED_REPORT, index=False)


agg = removed_rows.groupby("AimoScore").agg(
    count_removed=("AimoScore", "size"),
    mean_estimated=("EstimatedScore", "mean"),
    median_estimated=("EstimatedScore", "median"),
    mean_diff=("Diff", "mean")
).reset_index()

# Save kept rows to Excel via pandas (openpyxl engine creates file)
kept_rows.to_excel(OUTPUT_FILE, index=False, engine="openpyxl")

# Adjust column widths using openpyxl
wb = load_workbook(OUTPUT_FILE)
ws = wb.active

# Strategy: set each column width to max(length of header, max cell string length) capped to a sensible range
min_width = 15
max_width = 60

for i, col in enumerate(kept_rows.columns, start=1):
    col_letter = get_column_letter(i)
    # header length
    max_len = len(str(col))
    # check cell values in column
    for cell in ws[col_letter]:
        if cell.value is not None:
            cell_len = len(str(cell.value))
            if cell_len > max_len:
                max_len = cell_len
    # set width with caps and small padding
    adjusted_width = min(max(max_len + 2, min_width), max_width)
    ws.column_dimensions[col_letter].width = adjusted_width

wb.save(OUTPUT_FILE)

# Save reports
removed_rows.to_csv(REMOVED_REPORT, index=False)
agg.to_csv("AimoScore_removed_agg.csv", index=False)

print(f"Original rows: {len(df)}")
print(f"Removed rows: {len(removed_rows)} (duplicated AimoScore values: {len(duplicated_values)})")
print(f"Kept rows: {len(kept_rows)}")
